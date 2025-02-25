import openpyxl
import os
import json

def get_start_cells_for_sheet(sheet_name)
    sheet_start_cells = {
        "(f) Liquids":[(33,3),(1189,3),(1698,3)], 
        # Row 33, Column 3, etc.
        "(n) Solids":[(26,8),(79,4)]
        # Row 26, Column 8, etc.
    }

    # Return the corresponding start cells(row,rolumn) tuples, or default to [(33,3)] if sheet name is not in the list
    return sheet_start_cells.get(sheet_name, [(33,3)])

def find_table_start(sheet,start_row,start_col):
    # Find the row and column to start from
    return (start_row, start_col)

def scrape_excel_workbook(file_path):
    # Load the workbook
    workbook = openpyxl.load_workbook(file_path)
    # Create a dictionary to store the data
    scraped_data = {}
    # Iterate through the sheets
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        # Get the list of start rows and columns for the sheet
        start_cells = get_start_cells_for_sheet(sheet_name)
        # Initialize a list to store data for the current sheet
        sheet_data =[]
        # Iterate over each start cell (row,col) for the current sheet
        for start_row, start_col in start_cells:
            # Find the header in the sheet starting from the defined row (start_row) and column (start_col)
            header = None
            for row in sheet.iter_rows(min_row=start_row, min_col=start_col, max_col=sheet.max_column):
                if row[0].value is not None:
                    header = [cell.value for cell in row if cell.value is not None]
                    break

            # If the header is found, start collecting the data from the next row
            if header:
                # Iterate over the rows starting from the row after the header
                for row in sheet.iter_rows(min_row=start_row+1, min_col=start_col, max_col=sheet.max_column, values_only=True):
                    # Only collect rows that are not entirely empty
                    if any(cell is not None for cell in row):
                        row_dict = {header[i]: row[i] if i < len(row) else None for i in range(len(header))}
                        sheet_data.append(row_dict)
        # Store the data for the current sheet in the dictionary
        scraped_data[sheet_name] = sheet_data
    return scraped_data

def export_to_json(scraped_data, output_file):
    with open(output_file, 'w', encoding='utf-8') as json_file:
        json.dump(scraped_data, json_file, ensure_ascii=False, indent=4)

if __name__ == "__main__":
    file_path = "path/to/your/excel/file.xlsx"
    output_file = "path/to/output/file.json"
    scraped_data = scrape_excel_workbook(file_path)
    export_to_json(scraped_data, output_file)

print("Data scraped and exported to JSON successfully!")